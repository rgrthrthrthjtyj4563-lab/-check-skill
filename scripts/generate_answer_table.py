#!/usr/bin/env python3
"""Generate a single revised questionnaire answer-option wide table.

This script is intentionally conservative. It does not rewrite medical report
content. It treats an xlsx questionnaire detail file as source distribution,
removes metadata columns, samples answers by original option frequency, and
repairs rows that violate hard single-respondent logic rules. The final
artifact is one workbook with one sheet only.
"""

from __future__ import annotations

import argparse
import json
import random
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


META_COLUMNS = {
    "问卷名称",
    "品种",
    "被问卷人姓名",
    "问卷日期",
    "手机版本",
    "openid",
    "open id",
    "手机号",
    "手机号码",
    "地址",
    "备注",
    "受访者编号",
    "respondent_id",
    "respondent id",
}


NEGATIVE_PATTERNS = [
    "不用药",
    "几乎不用药",
    "未使用",
    "没使用",
    "不使用",
    "未听说",
    "不了解",
    "不清楚",
    "不确定",
    "无法评价",
    "无相关经历",
    "不适用",
]

USAGE_EVAL_QUESTION_PATTERNS = [
    "使用清咽滴丸",
    "购买清咽滴丸",
    "推荐清咽滴丸",
    "医保报销",
    "价格敏感度",
]

POSITIVE_OR_EVALUATIVE_PATTERNS = [
    "起效",
    "缓解",
    "效果",
    "便利",
    "方便",
    "推荐",
    "满意",
    "认可",
    "每次都使用医保",
    "偶尔使用医保",
    "当前价格合理",
    "促销活动时才购买",
    "疗效好",
]

QUESTION_PRIORITY = {
    "prerequisite": 0,
    "behavior": 1,
    "insurance": 2,
    "price": 3,
    "experience": 4,
    "attitude": 5,
    "channel": 6,
    "info": 7,
    "other": 8,
}


@dataclass
class Question:
    number: int
    text: str
    options: OrderedDict[str, int]


def normalize_header(value: object) -> str:
    return str(value or "").strip()


def is_meta_column(header: str) -> bool:
    normalized = header.strip().lower()
    if normalized in META_COLUMNS:
        return True
    return any(token in normalized for token in ["openid", "open id", "手机号", "手机号码"])


def parse_sample_size(prompt: str) -> int | None:
    patterns = [
        r"(?:样本量\s*)?N\s*[=:：]\s*(\d+)",
        r"生成\s*(\d+)\s*(?:个)?人",
        r"(\d+)\s*个人",
        r"受访者\s*(\d+)\s*人?",
        r"(\d+)\s*份问卷",
        r"样本量\s*(\d+)",
        r"(\d+)\s*个样本",
    ]
    for pattern in patterns:
        match = re.search(pattern, prompt, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))
    return None


def read_questions(path: Path) -> tuple[list[str], list[Question], dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [normalize_header(v) for v in next(rows)]
    meta_values: dict[str, str] = {}
    raw_rows = list(rows)

    if raw_rows:
        first = raw_rows[0]
        for idx, header in enumerate(headers):
            if header and is_meta_column(header) and idx < len(first) and first[idx] is not None:
                meta_values[header] = str(first[idx]).strip()

    meta_columns: list[str] = []
    questions: list[Question] = []
    for idx, header in enumerate(headers):
        if not header:
            continue
        if is_meta_column(header):
            meta_columns.append(header)
            continue
        counts: OrderedDict[str, int] = OrderedDict()
        for row in raw_rows:
            if idx >= len(row):
                continue
            value = row[idx]
            if value is None:
                continue
            option = str(value).strip()
            if not option:
                continue
            counts[option] = counts.get(option, 0) + 1
        if counts:
            questions.append(Question(len(questions) + 1, header, counts))
    return meta_columns, questions, meta_values


def weighted_choice(options: OrderedDict[str, int]) -> str:
    values = list(options.keys())
    weights = list(options.values())
    return random.choices(values, weights=weights, k=1)[0]


def contains_any(text: str, patterns: Iterable[str]) -> bool:
    return any(pattern in text for pattern in patterns)


def classify_question(question_text: str) -> str:
    text = question_text.lower()
    if contains_any(text, ["医保"]):
        return "insurance"
    if contains_any(text, ["推荐", "复购", "认可", "是否会向", "是否愿意"]):
        return "attitude"
    if contains_any(text, ["症状改善", "疗效", "效果", "满意", "安全", "便利", "依从", "使用"]):
        if contains_any(text, ["使用"]):
            return "experience"
        return "experience"
    if contains_any(text, ["用药频率", "管理方式", "是否使用", "是否听说", "是否了解"]):
        return "prerequisite"
    if contains_any(text, ["价格", "促销"]):
        return "price"
    if contains_any(text, ["渠道", "药店", "医院", "电商", "购买"]):
        return "channel"
    if contains_any(text, ["信息", "广告", "科普", "来源"]):
        return "info"
    return "other"


def classify_option(option_text: str) -> set[str]:
    text = option_text.lower()
    tags: set[str] = set()

    if contains_any(text, ["不确定", "不清楚", "不知道", "无法评价", "无法判断", "一般"]):
        tags.add("neutral")
    if contains_any(text, ["未使用", "没使用", "未听说", "不了解", "无相关经历", "不适用", "不用药靠自然缓解"]):
        tags.add("no_experience")
    if contains_any(text, ["不用药靠自然缓解", "几乎不用药", "半年1次", "半年 1 次"]):
        tags.add("low_frequency")
    if contains_any(text, ["每月至少", "长期", "规律", "定期", "每次都"]):
        tags.add("high_frequency")
    if contains_any(text, ["不使用医保", "全额自费"]):
        tags.add("insurance_not_used")
    if contains_any(text, ["每次都使用医保", "偶尔使用医保"]):
        tags.add("insurance_used")
    if contains_any(text, ["价格非常敏感", "只选低价", "促销活动时才购买"]):
        tags.add("price_sensitive")
    if contains_any(text, ["价格不是主要", "只要疗效好", "价格不是"]):
        tags.add("price_not_sensitive")

    if contains_any(text, ["不会推荐", "几乎没有效果", "无效", "不满意", "不喜欢", "不认可", "不方便", "不好"]):
        tags.add("negative")
    if contains_any(text, ["肯定会推荐", "起效快", "明显", "非常满意", "长期规律"]):
        tags.add("strong_positive")
        tags.add("positive")
    elif contains_any(text, ["可能会推荐", "有一定缓解", "缓解", "便利", "方便", "满意", "认可", "合理可接受", "疗效好"]):
        tags.add("positive")

    if not tags:
        tags.add("neutral")
    return tags


def question_by_type(questions: list[Question], qtype: str) -> list[Question]:
    return [question for question in questions if classify_question(question.text) == qtype]


def pick_option_by_tags(question: Question, preferred_tags: Iterable[str], fallback_tags: Iterable[str] = ()) -> str:
    preferred = set(preferred_tags)
    fallback = set(fallback_tags)
    for option in question.options:
        tags = classify_option(option)
        if preferred & tags:
            return option
    for option in question.options:
        tags = classify_option(option)
        if fallback & tags:
            return option
    return next(iter(question.options))


def is_option_allowed(row: dict[str, str], questions: list[Question], question: Question, option: str) -> bool:
    candidate = dict(row)
    candidate[question.text] = option
    return not validate_row(candidate, questions)


def validate_row(row: dict[str, str], questions: list[Question]) -> list[dict[str, str]]:
    conflicts: list[dict[str, str]] = []
    q_by_text = {q.text: q for q in questions}

    for q_text, answer in row.items():
        q = q_by_text[q_text]
        qtype = classify_question(q_text)
        tags = classify_option(answer)

        if tags & {"no_experience"} or (qtype == "prerequisite" and "neutral" in tags):
            for later_text, later_answer in row.items():
                later_q = q_by_text[later_text]
                later_type = classify_question(later_text)
                later_tags = classify_option(later_answer)
                if later_q.number <= q.number:
                    continue
                if later_type in {"experience", "attitude"} and later_tags & {"positive", "strong_positive", "negative"}:
                    conflicts.append(
                        {
                            "前置题号及选项": f"Q{q.number} {answer}",
                            "后续题号及选项": f"Q{later_q.number} {later_answer}",
                            "冲突类型": "前提不成立",
                            "冲突原因": "前置答案否定使用/认知前提，后续答案给出了明确体验或态度评价。",
                            "风险等级": "高",
                            "修改建议": "保留前提题，后续体验或态度题改为中性/无法评价。",
                        }
                    )

        if tags & {"low_frequency", "no_experience"}:
            for later_text, later_answer in row.items():
                later_q = q_by_text[later_text]
                later_tags = classify_option(later_answer)
                if later_q.number <= q.number:
                    continue
                if later_tags & {"high_frequency", "insurance_used"}:
                    conflicts.append(
                        {
                            "前置题号及选项": f"Q{q.number} {answer}",
                            "后续题号及选项": f"Q{later_q.number} {later_answer}",
                            "冲突类型": "行为频率冲突",
                            "冲突原因": "低频/不用药答案不能与长期规律、每次医保等强行为答案共存。",
                            "风险等级": "高",
                            "修改建议": "保留前置行为频率，修正后续强行为答案。",
                        }
                    )

        if tags & {"insurance_not_used"}:
            for other_text, other_answer in row.items():
                other_q = q_by_text[other_text]
                if other_text == q_text:
                    continue
                if classify_option(other_answer) & {"insurance_used"}:
                    conflicts.append(
                        {
                            "前置题号及选项": f"Q{q.number} {answer}",
                            "后续题号及选项": f"Q{other_q.number} {other_answer}",
                            "冲突类型": "医保使用互斥",
                            "冲突原因": "同一受访者不能同时选择不使用医保和使用医保报销。",
                            "风险等级": "高",
                            "修改建议": "保留一种医保使用状态。",
                        }
                    )

        if qtype == "price" and tags & {"price_sensitive"}:
            for other_text, other_answer in row.items():
                other_q = q_by_text[other_text]
                if other_text == q_text:
                    continue
                if classify_question(other_text) == "price" and classify_option(other_answer) & {"price_not_sensitive"}:
                    conflicts.append(
                        {
                            "前置题号及选项": f"Q{q.number} {answer}",
                            "后续题号及选项": f"Q{other_q.number} {other_answer}",
                            "冲突类型": "价格态度冲突",
                            "冲突原因": "高价格敏感不能与价格不是主要考虑在同一受访者中共存。",
                            "风险等级": "中",
                            "修改建议": "保留一个价格态度，另一个改为中性或同向选项。",
                        }
                    )

    experience_states = []
    attitude_states = []
    for q_text, answer in row.items():
        q = q_by_text[q_text]
        tags = classify_option(answer)
        qtype = classify_question(q_text)
        if qtype == "experience":
            experience_states.append((q, answer, tags))
        elif qtype == "attitude":
            attitude_states.append((q, answer, tags))

    has_positive_experience = any(tags & {"positive", "strong_positive"} for _, _, tags in experience_states)
    has_negative_experience = any("negative" in tags for _, _, tags in experience_states)
    for att_q, att_answer, att_tags in attitude_states:
        if has_positive_experience and "negative" in att_tags:
            source_q, source_answer, _ = next(
                (item for item in experience_states if item[2] & {"positive", "strong_positive"}),
                experience_states[0],
            )
            conflicts.append(
                {
                    "前置题号及选项": f"Q{source_q.number} {source_answer}",
                    "后续题号及选项": f"Q{att_q.number} {att_answer}",
                    "冲突类型": "体验与态度极性冲突",
                    "冲突原因": "正向体验评价不能与负向推荐/复购/认可态度共存。",
                    "风险等级": "高",
                    "修改建议": "保留体验评价，修正态度题为可能推荐/肯定推荐/中性。",
                }
            )
        if has_negative_experience and "strong_positive" in att_tags:
            source_q, source_answer, _ = next(
                (item for item in experience_states if "negative" in item[2]),
                experience_states[0],
            )
            conflicts.append(
                {
                    "前置题号及选项": f"Q{source_q.number} {source_answer}",
                    "后续题号及选项": f"Q{att_q.number} {att_answer}",
                    "冲突类型": "负向体验与强正向态度冲突",
                    "冲突原因": "负向体验评价不能与强推荐/强认可态度共存。",
                    "风险等级": "高",
                    "修改建议": "保留体验评价，修正态度题为不推荐或不确定。",
                }
            )
    return conflicts


def row_conflicts(row: dict[str, str], questions: list[Question]) -> list[dict[str, str]]:
    return validate_row(row, questions)


def safer_option(question: Question) -> str:
    options = list(question.options.keys())
    for option in options:
        if contains_any(option, ["不确定", "无法评价", "不清楚"]):
            return option
    for option in options:
        if not contains_any(option, POSITIVE_OR_EVALUATIVE_PATTERNS):
            return option
    return options[0]


def context_safe_option(row: dict[str, str], questions: list[Question], question: Question) -> str:
    qtype = classify_question(question.text)
    if qtype == "attitude":
        experiences = [
            classify_option(answer)
            for q_text, answer in row.items()
            if classify_question(q_text) == "experience"
        ]
        if any("negative" in tags for tags in experiences):
            return pick_option_by_tags(question, ["negative", "neutral"])
        if any(tags & {"positive", "strong_positive"} for tags in experiences):
            return pick_option_by_tags(question, ["positive", "strong_positive"], ["neutral"])
    if qtype == "experience":
        if any(classify_option(answer) & {"no_experience"} for answer in row.values()):
            return pick_option_by_tags(question, ["neutral"])
    return safer_option(question)


def repair_row(row: dict[str, str], questions: list[Question]) -> dict[str, str]:
    repaired = dict(row)
    q_by_number = {question.number: question for question in questions}
    for _ in range(12):
        conflicts = validate_row(repaired, questions)
        if not conflicts:
            return repaired
        for conflict in conflicts:
            match = re.match(r"Q(\d+)\s", conflict["后续题号及选项"])
            if not match:
                continue
            q_number = int(match.group(1))
            question = q_by_number[q_number]
            repaired[question.text] = context_safe_option(repaired, questions, question)
    return repaired


def constrained_weighted_choice(row: dict[str, str], questions: list[Question], question: Question) -> tuple[str, bool]:
    allowed: OrderedDict[str, int] = OrderedDict()
    for option, weight in question.options.items():
        if is_option_allowed(row, questions, question, option):
            allowed[option] = weight
    if allowed:
        return weighted_choice(allowed), False
    return context_safe_option(row, questions, question), True


def generate_answers(questions: list[Question], sample_size: int) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    answers: list[dict[str, str]] = []
    all_conflicts: list[dict[str, str]] = []
    adjusted_count = 0
    generation_order = sorted(questions, key=lambda question: (QUESTION_PRIORITY[classify_question(question.text)], question.number))
    for _ in range(sample_size):
        row: dict[str, str] = {}
        for question in generation_order:
            answer, adjusted = constrained_weighted_choice(row, questions, question)
            adjusted_count += int(adjusted)
            row[question.text] = answer
        row = repair_row(row, questions)
        conflicts = validate_row(row, questions)
        all_conflicts.extend(conflicts)
        answers.append(row)
    generate_answers.adjusted_count = adjusted_count
    return answers, all_conflicts


def write_outputs(
    output_dir: Path,
    source_path: Path,
    prompt: str,
    sample_size: int | None,
    meta_columns: list[str],
    questions: list[Question],
    meta_values: dict[str, str],
    answers: list[dict[str, str]],
    conflicts: list[dict[str, str]],
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = source_path.stem
    output_path = output_dir / f"{stem}_修改后问卷_答案选项宽表.xlsx"

    if sample_size is None:
        return {"output": "", "error": "缺少生成人数，无法生成答案选项表。"}

    wb = Workbook()
    ws = wb.active
    ws.title = "修改后问卷及答案表"
    headers = ["问卷日期"]
    headers.extend(f"问题{question.number}：{question.text}" for question in questions)
    headers.extend(["手机版本", "OpenID"])
    ws.append(headers)

    start = date.today()
    for row in answers:
        values = [(start - timedelta(days=random.randint(0, 30))).isoformat()]
        values.extend(row[question.text] for question in questions)
        values.extend(["", ""])
        ws.append(values)

    wb.save(output_path)
    return {"output": str(output_path)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--questionnaire", required=True, type=Path)
    parser.add_argument("--prompt", required=True)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--seed", type=int, default=20240521)
    args = parser.parse_args()

    random.seed(args.seed)
    sample_size = parse_sample_size(args.prompt)
    meta_columns, questions, meta_values = read_questions(args.questionnaire)
    answers: list[dict[str, str]] = []
    conflicts: list[dict[str, str]] = []
    if sample_size is not None:
        answers, conflicts = generate_answers(questions, sample_size)

    outputs = write_outputs(
        args.output_dir,
        args.questionnaire,
        args.prompt,
        sample_size,
        meta_columns,
        questions,
        meta_values,
        answers,
        conflicts,
    )
    print(
        json.dumps(
            {
                "sample_size": sample_size,
                "effective_question_count": len(questions),
                "meta_columns": meta_columns,
                "adjusted_option_count": getattr(generate_answers, "adjusted_count", 0),
                "conflict_count_after_generation": len(conflicts),
                **outputs,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 1 if sample_size is None or conflicts else 0


if __name__ == "__main__":
    raise SystemExit(main())
